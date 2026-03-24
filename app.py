from datetime import datetime
from pathlib import Path

from flask import Flask, render_template
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "data.xlsx"

app = Flask(__name__)


def normalize_label(value: str) -> str:
    return "".join(character.lower() for character in value if character.isalnum())


def find_matching_column(columns: list[str], aliases: list[str]) -> str | None:
    normalized_columns = {normalize_label(column): column for column in columns if column}
    for alias in aliases:
        match = normalized_columns.get(normalize_label(alias))
        if match:
            return match
    return None


def build_dashboard_context(columns: list[str], rows: list[dict]) -> dict:
    status_column = find_matching_column(columns, ["Status", "Project Status"])
    name_column = find_matching_column(columns, ["Project Name", "Name", "Project"])
    tools_column = find_matching_column(columns, ["Tools Used", "Tools", "Platform", "Platforms"])
    description_column = find_matching_column(columns, ["Description", "Summary", "Details"])

    completed_count = 0
    in_progress_count = 0
    dashboard_rows = []

    for row in rows:
        raw_status = str(row.get(status_column, "")).strip() if status_column else ""
        normalized_status = normalize_label(raw_status)
        if normalized_status in {"completed", "complete", "done"}:
            completed_count += 1
        elif normalized_status in {"inprogress", "progress", "active", "ongoing"}:
            in_progress_count += 1

        dashboard_rows.append(
            {
                "project_name": row.get(name_column, "") if name_column else "",
                "tools_used": row.get(tools_column, "") if tools_column else "",
                "description": row.get(description_column, "") if description_column else "",
                "status": raw_status,
            }
        )

    return {
        "total_projects": len(rows),
        "in_progress_count": in_progress_count,
        "completed_count": completed_count,
        "dashboard_rows": dashboard_rows,
        "has_dashboard_columns": all([name_column, tools_column, description_column, status_column]),
    }


def load_excel_data() -> tuple[list[str], list[dict], str | None]:
    if not EXCEL_FILE.exists():
        return [], [], f"Excel file not found: {EXCEL_FILE.name}"

    try:
        workbook = load_workbook(EXCEL_FILE, data_only=True)
        worksheet = workbook.active
        values = list(worksheet.iter_rows(values_only=True))
        if not values:
            return [], [], None

        header_row = values[0]
        columns = [str(cell).strip() if cell is not None else "" for cell in header_row]
        if not any(columns):
            return [], [], "The first row of the Excel file must contain column names."

        rows = []
        for excel_row in values[1:]:
            row = {}
            for index, column in enumerate(columns):
                if not column:
                    continue
                cell_value = excel_row[index] if index < len(excel_row) else ""
                row[column] = "" if cell_value is None else cell_value
            if any(value != "" for value in row.values()):
                rows.append(row)

        return columns, rows, None
    except Exception as exc:
        return [], [], f"Could not read {EXCEL_FILE.name}: {exc}"
    finally:
        if "workbook" in locals():
            workbook.close()


@app.after_request
def add_no_cache_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.route("/")
def index():
    columns, rows, error = load_excel_data()
    dashboard = build_dashboard_context(columns, rows)
    file_last_updated = None
    if EXCEL_FILE.exists():
        file_last_updated = datetime.fromtimestamp(EXCEL_FILE.stat().st_mtime).strftime("%b %d, %Y %I:%M:%S %p")

    return render_template(
        "index.html",
        columns=columns,
        rows=rows,
        error=error,
        excel_file=EXCEL_FILE.name,
        year=datetime.now().year,
        dashboard_loaded_at=datetime.now().strftime("%b %d, %Y %I:%M:%S %p"),
        file_last_updated=file_last_updated,
        **dashboard,
    )


if __name__ == "__main__":
    app.run(debug=True)
